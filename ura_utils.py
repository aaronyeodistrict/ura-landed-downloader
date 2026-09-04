"""
URA scraping + Excel utilities — adapted from V1 desktop app for web/server use.
No tkinter. _write_excel returns BytesIO instead of saving to disk.
"""

import subprocess, json, re, csv, io, urllib.parse, html as _html, tempfile, os
from datetime import datetime
from pathlib import Path

# ── Constants ──────────────────────────────────────────────────────────────
MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

POSTAL_DISTRICTS = [
    "D01 / Boat Quay, Raffles Place, Marina",
    "D02 / Anson, Tanjong Pagar",
    "D03 / Queenstown, Tiong Bahru",
    "D04 / Telok Blangah, Harbourfront",
    "D05 / Pasir Panjang, Hong Leong Garden, Clementi New Town",
    "D06 / High Street, Beach Road",
    "D07 / Middle Road, Golden Mile",
    "D08 / Little India",
    "D09 / Orchard, Cairnhill, River Valley",
    "D10 / Ardmore, Bukit Timah, Holland Road, Tanglin",
    "D11 / Watten Estate, Novena, Thomson",
    "D12 / Balestier, Toa Payoh, Serangoon",
    "D13 / Macpherson, Braddell",
    "D14 / Geylang, Eunos",
    "D15 / Katong, Joo Chiat, Amber Road",
    "D16 / Bedok, Upper East Coast, Eastwood, Kew Drive",
    "D17 / Loyang, Changi",
    "D18 / Tampines, Pasir Ris",
    "D19 / Serangoon Garden, Hougang, Punggol",
    "D20 / Bishan, Ang Mo Kio",
    "D21 / Upper Bukit Timah, Clementi Park, Ulu Pandan",
    "D22 / Jurong",
    "D23 / Hillview, Dairy Farm, Bukit Panjang, Choa Chu Kang",
    "D24 / Lim Chu Kang, Tengah",
    "D25 / Kranji, Woodgrove",
    "D26 / Upper Thomson, Springleaf",
    "D27 / Yishun, Sembawang",
    "D28 / Seletar",
]

LANDED_TYPES = {
    "Detached", "Semi-detached", "Terrace",
    "Detached House", "Semi-Detached House", "Terrace House",
}

_now = datetime.now()
URA_EARLIEST_YEAR  = _now.year - 5
URA_EARLIEST_MONTH = 1
YEAR_RANGE = list(range(URA_EARLIEST_YEAR, _now.year + 1))

COL_MAP = [
    ("Project Name", "Project Name"),
    ("Street",       "Street Name"),
    ("Area",         "Area (sqft)"),
    ("Per Square Foot", "Unit Price ($ PSF)"),
    ("Price",        "Transacted Price ($)"),
    ("Tenure",       "Tenure"),
    ("Date",         "Sale Date"),
    ("Property Type","Property Type"),
]

# ── Network helpers ────────────────────────────────────────────────────────
def _curl(url, headers, timeout=90):
    cmd = ["curl","-sL","--max-time",str(timeout),"--connect-timeout","20","-k",
           "--http1.1","--retry","2","--retry-delay","3",
           "-A","Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/125.0 Safari/537.36"]
    for k, v in headers.items():
        cmd += ["-H", f"{k}: {v}"]
    cmd.append(url)
    result = subprocess.run(cmd, capture_output=True, timeout=timeout + 30)
    if result.returncode != 0:
        code   = result.returncode
        detail = result.stderr.decode("utf-8", "replace").strip() or "no response"
        hint   = " (URA server not responding — check internet)" if code == 28 else ""
        raise RuntimeError(f"Network error (curl {code}): {detail}{hint}")
    raw = result.stdout
    if not raw.strip():
        raise RuntimeError("Empty response from URA server.")
    try:
        return json.loads(raw.decode("utf-8", errors="replace"))
    except json.JSONDecodeError:
        preview = raw[:200].decode("utf-8", "replace").replace("\n", " ")
        raise RuntimeError(f"Unexpected response (not JSON): {preview}")


def _curl_raw(url, extra_args=None, timeout=15):
    cmd = ["curl","-sL","--max-time",str(timeout),"-k",
           "-A","Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 Chrome/125.0 Safari/537.36"]
    if extra_args:
        cmd += extra_args
    cmd.append(url)
    r = subprocess.run(cmd, capture_output=True, timeout=timeout + 5)
    return r.stdout, r.stderr.decode("utf-8", "replace"), r.returncode


def _ura_website_csrf_and_cookies():
    cookie_file = os.path.join(tempfile.gettempdir(), f"ura_cookies_{os.getpid()}.txt")
    base = "https://eservice.ura.gov.sg/property-market-information/pmiResidentialTransactionSearch"
    stdout, _, rc = _curl_raw(base, extra_args=["-c", cookie_file, "-D", "-"], timeout=15)
    raw  = stdout.decode("utf-8", errors="replace")
    csrf = None
    for pat in [
        r'<meta[^>]+name=["\']_csrf["\'][^>]+content=["\']([^"\']+)["\']',
        r'<meta[^>]+content=["\']([^"\']+)["\'][^>]+name=["\']_csrf["\']',
        r'_csrf.*?([0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12})',
    ]:
        m = re.search(pat, raw, re.IGNORECASE)
        if m:
            csrf = m.group(1)
            break
    if not csrf:
        raise RuntimeError("Could not get CSRF token from URA website. Check internet connection.")
    return csrf, cookie_file


# ── Project list (cached externally by Streamlit) ──────────────────────────
def get_all_project_names_website(progress_cb=None):
    """Scan all postal districts and return sorted list of landed project names."""
    base        = "https://eservice.ura.gov.sg/property-market-information/pmiResidentialTransactionSearch"
    dl_url_hard = "https://eservice.ura.gov.sg/property-market-information/pmiSearchResidentialTransactionDownload"
    now = datetime.now()

    def _enc(fields):
        return "&".join(
            f"{urllib.parse.quote(str(k), safe='[]')}={urllib.parse.quote(str(v), safe='')}"
            for k, v in fields.items()
        )

    def _district_names(district, csrf, cookie_file):
        loc_json = json.dumps(["postalDistrict", district], separators=(",", ":"))
        search_fields = {
            "resultPerPage": "20", "displayResult": "true", "displayResultHeader": "true",
            "loadAnalysis": "true", "displayAnalysis": "false", "displayChart": "true",
            "displayAnalysisFilters": "true", "dashboardDisplay": "false",
            "locationDetails": loc_json,
            "saleYearFrom": str(now.year - 3), "saleMonthFrom": "1",
            "saleYearTo": str(now.year), "saleMonthTo": str(now.month),
            "saleType": "3", "_saleType": "1", "propertyTypeGroupNo": "1", "_csrf": csrf,
        }
        c_srch = ["-H", f"Referer: {base}", "-H", "Origin: https://eservice.ura.gov.sg",
                  "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                  "-H", "X-Requested-With: XMLHttpRequest", "-b", cookie_file, "-c", cookie_file,
                  "-H", "Content-Type: application/x-www-form-urlencoded"]
        c_dl   = ["-H", f"Referer: {base}", "-H", "Origin: https://eservice.ura.gov.sg",
                  "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
                  "-b", cookie_file, "-c", cookie_file,
                  "-H", "Content-Type: application/x-www-form-urlencoded"]
        st1, _, _ = _curl_raw(base, extra_args=c_srch + ["-X", "POST", "-d", _enc(search_fields)], timeout=30)
        resp1 = st1.decode("utf-8", errors="replace")
        if "resultForm" not in resp1:
            return set()
        dl_fields = {}; actual_url = dl_url_hard
        fm = re.search(r'<form[^>]*id=["\']resultForm1["\'][^>]*>(.*?)</form>', resp1, re.DOTALL)
        if not fm:
            return set()
        am = (re.search(r'<form[^>]*id=["\']resultForm1["\'][^>]*action=["\']([^"\']+)["\']', resp1) or
              re.search(r'<form[^>]*action=["\']([^"\']+)["\'][^>]*id=["\']resultForm1["\']', resp1))
        if am:
            a = am.group(1)
            actual_url = a if a.startswith("http") else "https://eservice.ura.gov.sg" + a
        for inp in re.findall(r"<input[^>]+>", fm.group(1)):
            nm = re.search(r'\bname=["\']([^"\']+)["\']', inp)
            vl = re.search(r'\bvalue=["\']([^"\']*)["\']', inp)
            if nm:
                dl_fields[nm.group(1)] = _html.unescape(vl.group(1)) if vl else ""
        dl_fields["downloadType"] = "downloadCSV"; dl_fields["gotoPage"] = "1"
        if "_csrf" not in dl_fields:
            dl_fields["_csrf"] = csrf
        dl_body = _enc(dl_fields)
        for c in range(1, 18):
            dl_body += f"&selectColumn={c}"
        st2, _, _ = _curl_raw(actual_url, extra_args=c_dl + ["-X", "POST", "-d", dl_body, "-L", "--max-redirs", "5"], timeout=60)
        csv_text = st2.decode("utf-8-sig", errors="replace").lstrip("\r\n ")
        if "Project Name" not in csv_text[:500]:
            return set()
        names = set()
        reader = csv.DictReader(io.StringIO(csv_text))
        for rec in reader:
            if rec.get("Property Type", "").strip() in LANDED_TYPES:
                pname = rec.get("Project Name", "").strip()
                if pname:
                    names.add(pname)
        return names

    all_names = set()
    total = len(POSTAL_DISTRICTS)
    for i, district in enumerate(POSTAL_DISTRICTS):
        if progress_cb:
            try:
                progress_cb(i / total, f"Scanning district {i+1}/{total}: {district[:35]}...")
            except Exception:
                pass
        try:
            csrf, cookie_file = _ura_website_csrf_and_cookies()
            all_names.update(_district_names(district, csrf, cookie_file))
        except Exception:
            pass
    return sorted(all_names)


# ── Date helpers ───────────────────────────────────────────────────────────
def parse_date(mmyy):
    try:
        mm = int(mmyy[:2]); yy = int(mmyy[2:])
        return datetime(2000 + yy if yy < 100 else yy, mm, 1)
    except Exception:
        return None


def sqm_to_sqft(s):
    try:
        return int(float(s) * 10.7639 + 0.5)
    except Exception:
        return ""


def psm_to_psf(price, area):
    try:
        p = float(price); a = float(area)
        return round(p / (a * 10.7639)) if a > 0 else ""
    except Exception:
        return ""


def _parse_csv_results(csv_text, loc_filters, from_dt, to_dt, prop_types=None):
    loc_lowers   = [f.strip().lower() for f in (loc_filters or []) if f.strip()]
    allowed_types = prop_types if prop_types else LANDED_TYPES
    rows = []
    reader = csv.DictReader(io.StringIO(csv_text))
    for rec in reader:
        pname     = rec.get("Project Name", "").strip()
        prop_type = rec.get("Property Type", "").strip()
        if prop_type not in LANDED_TYPES:
            continue
        if prop_type not in allowed_types:
            continue
        if loc_lowers and not any(ll in pname.lower() for ll in loc_lowers):
            continue
        sale_date_str = rec.get("Sale Date", "").strip()
        try:
            sd = datetime.strptime(sale_date_str, "%b-%y")
        except Exception:
            try:
                sd = datetime.strptime(sale_date_str, "%b-%Y")
            except Exception:
                sd = None
        if from_dt and sd and sd < from_dt:
            continue
        if to_dt and sd and sd > to_dt:
            continue

        def _int(s):
            try:
                return int(float(str(s).replace(",", "").strip()) + 0.5)
            except Exception:
                return s

        rows.append({
            "Project Name":        pname,
            "Street Name":         rec.get("Street Name", "").strip().title(),
            "Area (sqft)":         _int(rec.get("Area (SQFT)", "")),
            "Transacted Price ($)": _int(rec.get("Transacted Price ($)", "")),
            "Unit Price ($ PSF)":  _int(rec.get("Unit Price ($ PSF)", "")),
            "Sale Date":           sd or sale_date_str,
            "Tenure":              rec.get("Tenure", "").strip(),
            "Property Type":       prop_type,
        })
    return rows


# ── Main data fetcher ──────────────────────────────────────────────────────
def collect_rows_website(progress_cb, loc_filters=None, from_dt=None, to_dt=None,
                         mode="project", prop_types=None):
    base         = "https://eservice.ura.gov.sg/property-market-information/pmiResidentialTransactionSearch"
    download_url = "https://eservice.ura.gov.sg/property-market-information/pmiSearchResidentialTransactionDownload"

    progress_cb("Connecting to URA website...")
    csrf, cookie_file = _ura_website_csrf_and_cookies()

    now     = datetime.now()
    yr_from = from_dt.year  if from_dt else now.year - 5
    mo_from = from_dt.month if from_dt else 1
    yr_to   = to_dt.year    if to_dt   else now.year
    mo_to   = to_dt.month   if to_dt   else now.month

    def _enc(fields):
        return "&".join(
            f"{urllib.parse.quote(str(k), safe='[]')}={urllib.parse.quote(str(v), safe='')}"
            for k, v in fields.items()
        )

    common_search = [
        "-H", f"Referer: {base}", "-H", "Origin: https://eservice.ura.gov.sg",
        "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "-H", "X-Requested-With: XMLHttpRequest",
        "-b", cookie_file, "-c", cookie_file,
        "-H", "Content-Type: application/x-www-form-urlencoded",
    ]
    common_dl = [
        "-H", f"Referer: {base}", "-H", "Origin: https://eservice.ura.gov.sg",
        "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "-b", cookie_file, "-c", cookie_file,
        "-H", "Content-Type: application/x-www-form-urlencoded",
    ]

    projects = loc_filters if loc_filters else []
    if not projects:
        raise RuntimeError("No projects/districts selected.")

    all_rows = []
    for idx, item in enumerate(projects):
        progress_cb(f"Fetching {item} ({idx+1}/{len(projects)})...")
        loc_json = json.dumps(
            ["postalDistrict", item] if mode == "district" else ["projectName", item],
            separators=(",", ":"),
        )
        search_fields = {
            "resultPerPage": "20", "displayResult": "true", "displayResultHeader": "true",
            "loadAnalysis": "true", "displayAnalysis": "false", "displayChart": "true",
            "displayAnalysisFilters": "true", "dashboardDisplay": "false",
            "locationDetails": loc_json,
            "saleYearFrom": str(yr_from), "saleMonthFrom": str(mo_from),
            "saleYearTo": str(yr_to), "saleMonthTo": str(mo_to),
            "saleType": "3", "_saleType": "1", "_csrf": csrf,
        }
        if mode == "district":
            search_fields["propertyTypeGroupNo"] = "1"

        st1, _, _ = _curl_raw(
            base,
            extra_args=common_search + ["-X", "POST", "-d", _enc(search_fields)],
            timeout=30,
        )
        resp1 = st1.decode("utf-8", errors="replace")
        if "resultForm" not in resp1:
            continue

        dl_fields = {}; actual_dl_url = download_url
        form_m = re.search(r'<form[^>]*id=["\']resultForm1["\'][^>]*>(.*?)</form>', resp1, re.DOTALL)
        if form_m:
            action_m = (
                re.search(r'<form[^>]*id=["\']resultForm1["\'][^>]*action=["\']([^"\']+)["\']', resp1) or
                re.search(r'<form[^>]*action=["\']([^"\']+)["\'][^>]*id=["\']resultForm1["\']', resp1)
            )
            if action_m:
                a = action_m.group(1)
                actual_dl_url = a if a.startswith("http") else "https://eservice.ura.gov.sg" + a
            for inp in re.findall(r"<input[^>]+>", form_m.group(1)):
                nm = re.search(r'\bname=["\']([^"\']+)["\']', inp)
                vl = re.search(r'\bvalue=["\']([^"\']*)["\']', inp)
                if nm:
                    dl_fields[nm.group(1)] = _html.unescape(vl.group(1)) if vl else ""
        else:
            continue

        dl_fields["downloadType"] = "downloadCSV"; dl_fields["gotoPage"] = "1"
        if "_csrf" not in dl_fields:
            dl_fields["_csrf"] = csrf
        dl_body = _enc(dl_fields)
        for col_num in range(1, 18):
            dl_body += f"&selectColumn={col_num}"

        st2, _, _ = _curl_raw(
            actual_dl_url,
            extra_args=common_dl + ["-X", "POST", "-d", dl_body, "-L", "--max-redirs", "5"],
            timeout=60,
        )
        csv_text = st2.decode("utf-8-sig", errors="replace").lstrip("\r\n ")
        if "Project Name" not in csv_text[:500] and not csv_text.startswith('"Project'):
            continue

        rows = _parse_csv_results(csv_text, None, from_dt, to_dt, prop_types=prop_types)
        all_rows.extend(rows)

    return all_rows


# ── Excel export (returns bytes for st.download_button) ───────────────────
def write_excel_to_bytes(rows, loc_filters):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    def _sort_key(r):
        sd = r.get("Sale Date")
        if isinstance(sd, datetime): return sd
        try: return datetime.strptime(str(sd), "%b-%y")
        except:
            try: return datetime.strptime(str(sd), "%b-%Y")
            except: return datetime.min

    rows      = sorted(rows, key=_sort_key, reverse=True)
    headers   = [c[0] for c in COL_MAP]
    data_keys = [c[1] for c in COL_MAP]
    hf        = PatternFill("solid", fgColor="1F4E79")
    hfont     = Font(bold=True, color="FFFFFF", size=10)
    thin      = Side(style="thin", color="B8CCE4")
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    COL_FMT   = {3: '#,##0" sqft"', 4: "$#,##0", 5: "$#,##0", 7: "mmm-yy"}
    alt       = PatternFill("solid", fgColor="D9E8F5")
    col_widths = [28, 26, 12, 20, 16, 12, 32, 20]

    def _write_sheet(ws, sheet_rows):
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = hf; c.font = hfont
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        for ri, row in enumerate(sheet_rows, 2):
            fill = alt if ri % 2 != 0 else None
            for col, (h, dk) in enumerate(zip(headers, data_keys), 1):
                val = row.get(dk, "")
                c = ws.cell(row=ri, column=col, value=val)
                c.border = border
                c.alignment = Alignment(vertical="center", horizontal="center")
                if fill: c.fill = fill
                if col in COL_FMT: c.number_format = COL_FMT[col]
        for col, w in enumerate(col_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "All Transactions"
    _write_sheet(ws, rows)

    for sheet_name, prop_type in [
        ("Detached House",      "Detached House"),
        ("Semi-Detached House", "Semi-Detached House"),
        ("Terrace House",       "Terrace House"),
    ]:
        filtered = [r for r in rows if r.get("Property Type", "") == prop_type]
        ws2 = wb.create_sheet(title=sheet_name)
        _write_sheet(ws2, filtered)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
